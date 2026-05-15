import openpyxl

# ==========================================
# 1. DEFINISI FUNGSI KEANGGOTAAN (FUZZIFICATION) // untuk harga dan keanggotaan pelayanan
# ==========================================

def fuzzify_pelayanan(x):
    """Menghitung derajat keanggotaan untuk Pelayanan [1-100]"""
    """menggunakan fungsi keanggotaan segitiga untuk Biasa, dan fungsi keanggotaan trapezoid untuk Buruk dan Bagus"""
    """Buruk: 0-50, Biasa: 40-80, Bagus: 70-100"""

    buruk = biasa = bagus = 0 # Inisialisasi derajat keanggotaan
    
    # Buruk (Turun): 0-50
    if x <= 30: buruk = 1
    elif 30 < x < 50: buruk = (50 - x) / (50 - 30)
    
    # Biasa (Segitiga): 40-80
    if 40 < x <= 60: biasa = (x - 40) / (60 - 40)
    elif 60 < x < 80: biasa = (80 - x) / (80 - 60)
    
    # Bagus (Naik): 70-100
    if 70 < x <= 90: bagus = (x - 70) / (90 - 70)
    elif x > 90: bagus = 1
    
    return {"buruk": buruk, "biasa": biasa, "bagus": bagus}

def fuzzify_harga(x):
    """Menghitung derajat keanggotaan untuk Harga [25k-55k]"""
    """menggunakan metode yang sama dengan pelayanan, tetapi dengan rentang yang berbeda"""
    """murah (25k-35k), sedang (30k-45k), mahal (40k-55k)"""
    murah = sedang = mahal = 0
    
    # Murah (Turun): 25k-35k
    if x <= 25000: murah = 1
    elif 25000 < x < 35000: murah = (35000 - x) / (35000 - 25000)
    
    # Sedang (Segitiga): 30k-45k
    if 30000 < x <= 37500: sedang = (x - 30000) / (37500 - 30000)
    elif 37500 < x < 45000: sedang = (45000 - x) / (45000 - 37500)
    
    # Mahal (Naik): 40k-55k
    if 40000 < x <= 50000: mahal = (x - 40000) / (50000 - 40000)
    elif x > 50000: mahal = 1
    
    return {"murah": murah, "sedang": sedang, "mahal": mahal} # Fungsi untuk menghitung derajat keanggotaan harga

# ==========================================
# 2. MESIN INFERENSI & DEFUZZIFICATION (SUGENO) // Menghitung skor kelayakan berdasarkan aturan fuzzy
# ==========================================


def get_kelayakan_score(pelayanan_val, harga_val):
    
    """Menghitung skor kelayakan menggunakan metode Sugeno berdasarkan nilai pelayanan dan harga"""
    """Output Sugeno: Rendah = 50, Sedang = 75, Tinggi = 100"""
    """metode yang digunakan adalah weighted average, dimana setiap aturan memiliki bobot yang dihitung dari derajat keanggotaan input dan nilai output yang telah ditentukan"""


    p = fuzzify_pelayanan(pelayanan_val)
    h = fuzzify_harga(harga_val)
    
    # Output Sugeno (Singletons)
    # Rendah = 50, Sedang = 75, Tinggi = 90, Sangat Tinggi = 100
    R, S, T, ST = 50, 75, 90, 100 # Definisi nilai output

    rules = []
    
    # Kumpulan Aturan (Inference Rules)
    # IF Pelayanan ... AND Harga ... THEN Kelayakan ...
    # Kita menggunakan MIN untuk operasi AND
    
    
    # 1. Pelayanan Buruk
    rules.append((min(p['buruk'], h['murah']), S))   # Buruk + Murah = Sedang
    rules.append((min(p['buruk'], h['sedang']), R))  # Buruk + Sedang = Rendah
    rules.append((min(p['buruk'], h['mahal']), R))   # Buruk + Mahal = Rendah
    
    # 2. Pelayanan Biasa
    rules.append((min(p['biasa'], h['murah']), T))   # Biasa + Murah = Tinggi
    rules.append((min(p['biasa'], h['sedang']), S))  # Biasa + Sedang = Sedang
    rules.append((min(p['biasa'], h['mahal']), R))   # Biasa + Mahal = Rendah
    
    # 3. Pelayanan Bagus
    rules.append((min(p['bagus'], h['murah']), ST))  # Bagus + Murah = Sangat Tinggi (ST)
    rules.append((min(p['bagus'], h['sedang']), T))  # Bagus + Sedang = Tinggi
    rules.append((min(p['bagus'], h['mahal']), S))   # Bagus + Mahal = Sedang
    
    # Defuzzification: Weighted Average
    numerator = sum(weight * score for weight, score in rules)  # Menghitung pembilang
    denominator = sum(weight for weight, score in rules)  # Menghitung penyebut

    return numerator / denominator if denominator > 0 else 0 # Menghindari pembagian dengan nol

# ==========================================
# 3. PROSES UTAMA // Membaca data, menghitung skor, dan menyimpan hasil ranking
# ==========================================

def main():
    # Load data
    wb = openpyxl.load_workbook('restoran.xlsx')
    sheet = wb.active
    
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            score = get_kelayakan_score(row[1], row[2])
            data.append({
                'id': row[0], 
                'pelayanan': row[1], 
                'harga': row[2], 
                'total_score': score
            })

    # Urutkan berdasarkan skor tertinggi, pelayanan tertinggi, dan harga terendah
    data.sort(key=lambda x: (x['total_score'], x['pelayanan'], -x['harga']), reverse=True)

    # Simpan hasil Top 5
    wb_ranking = openpyxl.Workbook()
    sheet_ranking = wb_ranking.active
    sheet_ranking.title = "Ranking Restoran"
    
    headers = ["ID Restoran", "Kualitas Pelayanan", "Harga", "Skor Kelayakan", "Kesimpulan"]
    sheet_ranking.append(headers)
    
    for d in data[:5]:
        score = d['total_score']
        if score >= 95:
            kesimpulan = "Sangat Layak (Pilihan Utama)"
        elif score >= 85:
            kesimpulan = "Layak (Direkomendasikan)"
        elif score >= 75:
            kesimpulan = "Cukup Layak"
        else:
            kesimpulan = "Kurang Layak"
            
        sheet_ranking.append([d['id'], d['pelayanan'], d['harga'], round(score, 2), kesimpulan])
    
    wb_ranking.save('ranking.xlsx')
    print("Selesai! 5 Restoran terbaik telah disimpan di ranking.xlsx")

if __name__ == "__main__":
    main()
